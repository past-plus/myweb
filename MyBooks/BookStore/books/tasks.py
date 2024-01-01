# _*_ coding: utf-8 _*_
# @Time    : 2023/12/21 22:40
# @Author  : PASS
import datetime
import os
from time import sleep
from celery import shared_task
import logging

from books.models import Book

logger = logging.getLogger(__name__)


@shared_task
def task(x):
    logger.info('this is test task ')
    return x


@shared_task
def scheduletask():

    from openpyxl import Workbook
    import datetime
    from io import BytesIO
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '书籍月度报告'  # 给工作表1设置标题
    row_one = ['题目名称', '作者名称','出版时间']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = Book.objects.all()
    # print(all_obj)
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.title, obj.author,obj.publication_date]
        if obj_info[2].tzinfo is not None:
            # 如果是带有时区信息的datetime对象，则移除时区信息
            obj_info[2] = obj_info[2].replace(tzinfo=None)
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中

    now = datetime.datetime.now()
    timezone=now.strftime("%Y-%m-%d-%H:%M:%S")
    # timestamp = datetime.timestamp(now)
    file_excel=timezone.split('-')[1]
    file_name="E:\desktop\{}".format(file_excel+"-Month_Report.xls")
    if not os.path.exists(file_name):
        wb.save(file_name)
    # wb.save()
    logger.info('this is scheduletask '+timezone)
    return 'schedule task'

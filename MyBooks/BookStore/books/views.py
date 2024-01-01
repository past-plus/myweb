import csv

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from books.models import Book
from django import forms


from . import tasks
# Create your views here.


def runtask(request):
    content= {'200': 'run task test success!---'}
    return JsonResponse(content)


def runscheduletask(request):
    tasks.scheduletask.delay()
    content= {'200': 'success！'}
    return JsonResponse(content)

def export_excel(request):
    from openpyxl import Workbook
    import datetime
    from io import BytesIO
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    sheet1.title = '书籍月度报告'  # 给工作表1设置标题
    row_one = ['题目名称', '作者名称','出版时间']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    all_obj = Book.objects.all()
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.title, obj.author,obj.publication_date]
        if obj_info[2].tzinfo is not None:
            # 如果是带有时区信息的datetime对象，则移除时区信息
            obj_info[2] = obj_info[2].replace(tzinfo=None)
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S').split('-')[1]
    file_name = '%s_Month_Report.xls' % ctime  # 给文件名中添加日期时间
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    return response


class CSVUploadForm(forms.Form):
    file = forms.FileField()

    class Meta:
        model = Book
        fields = "__all__"


def book_list(request):
    # students = []
    # for i in range(10):
    #     students.append(Book(title=f"Student {i}", author=f'tom {i}'))
    # Book.objects.bulk_create(students)
    # queryset = Book.objects.all()
    # queryset.delete()

    # students = []
    # for i in range(20):
    #     students.append(Book(title=f"title {i}", author=f'author {i}',publication_date='2023-12-20'))
    # Book.objects.bulk_create(students)
    queryset = Book.objects.all()

    return render(request, 'book_list.html', {'queryset': queryset})


from openpyxl import load_workbook


def upload_csv(request):
    file_object = request.FILES.get('csv')
    # 1.获取用户上传的文件对象
    type_text = file_object.name.split('.')[1]
    # print(type_text)
    if type_text in ['xls', 'csv']:
        try:
            wb = load_workbook(file_object)
            sheet = wb.worksheets[0]
            # 3.循环获取每一行数据
            for row in sheet.iter_rows(min_row=2):  # 从第二行开始
                text = row[0].value
                author = row[1].value
                # print(text, author)
                exists = Book.objects.filter(title=text).exists()
                if not exists:
                    Book.objects.create(title=text, author=author)
        except:
            return HttpResponse('解析文本格式失败，请检查格式内容')
    else:
        return HttpResponse('当前非csv,xls格式，请重新选择文件')

    return redirect('/')


def upload_csv_(request):
    if request.method == 'POST':
        queryset = CSVUploadForm(request.POST, request.FILES)
        if queryset.is_valid():
            # 读取CSV文件并将数据存入数据库
            file = request.FILES.get('file')
            if file:
                rows = csv.reader(file.read().decode().splitlines())
                for row in rows:
                    _, created = Book.objects.get_or_create(
                        column1=row[0], column2=row[1], column3=row[2]
                    )
    else:  # Get
        queryset = CSVUploadForm()
    return render(request, 'book_list.html', {'queryset': queryset})
